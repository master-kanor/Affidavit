from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/projects/affidavit-f87b34c0/credentials.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
print('sheets:', ', '.join(wb.sheetnames))
for ws in wb.worksheets:
    print(f'[{ws.title}]')
    for row in ws.iter_rows(values_only=True):
        values = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if not values:
            continue
        labels = []
        for value in values:
            lowered = value.lower()
            if any(token in lowered for token in ('password', 'secret', 'token', 'key', 'credential', 'private')):
                labels.append(f'{value}=[REDACTED]')
            elif 'http' in lowered or '@' in value or value.lower() in {'owner', 'admin', 'user', 'supabase', 'cloudflare', 'github'}:
                labels.append(value)
            else:
                labels.append(f'{value[:80]}')
        print(' | '.join(labels))
